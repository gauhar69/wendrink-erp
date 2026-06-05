import sqlite3
import openpyxl
import uuid
from decimal import Decimal
from datetime import datetime, timezone
import os

DB_PATH = "wendrink.db"
PRODUCTS_XLSX = "продукты  .xlsx"
RECIPES_XLSX = "Рецепты.xlsx"

def main():
    if not os.path.exists(PRODUCTS_XLSX) or not os.path.exists(RECIPES_XLSX):
        print(f"❌ Ошибка: Убедитесь, что файлы '{PRODUCTS_XLSX}' и '{RECIPES_XLSX}' находятся в этой папке.")
        return

    print("=" * 60)
    print("  ИМПОРТ МЕНЮ И РЕЦЕПТОВ ИЗ EXCEL В БАЗУ ДАННЫХ")
    print("=" * 60)

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    try:
        # 1. Загружаем существующие продукты и ингредиенты из БД для сопоставления
        existing_products_by_code = {}
        for r in cur.execute("SELECT id, pos_code, name, sku FROM products").fetchall():
            if r[1] is not None:
                existing_products_by_code[int(r[1])] = {"id": r[0], "name": r[2], "sku": r[3]}

        ingredient_map = {}
        for r in cur.execute("SELECT id, name, unit FROM ingredients").fetchall():
            # Очищаем имя от лишних пробелов для точного совпадения
            ingredient_map[r[1].strip().lower()] = r[0]

        # 2. Читаем продукты из Excel
        print(f"\n1. Чтение '{PRODUCTS_XLSX}'...")
        wb_p = openpyxl.load_workbook(PRODUCTS_XLSX, data_only=True)
        ws_p = wb_p.active

        products_to_insert = []
        products_to_update = []
        
        # Пропускаем шапку
        rows_p = list(ws_p.iter_rows(values_only=True))
        header_p = rows_p[0]
        
        for idx, row in enumerate(rows_p[1:], start=2):
            if row[0] is None or str(row[0]).strip() == "":
                continue
            
            try:
                pos_code = int(row[0])
                name = str(row[1]).strip()
                price = Decimal(str(row[3])) if row[3] is not None else Decimal("0")
                category = str(row[2]).strip() if row[2] else None
            except Exception as e:
                print(f"  ⚠️ Ошибка чтения строки {idx} в продуктах: {e}. Пропускаем.")
                continue

            if pos_code in existing_products_by_code:
                # Обновляем существующий продукт
                products_to_update.append((name, float(price), category, pos_code))
                existing_products_by_code[pos_code]["name"] = name
            else:
                # Создаем новый продукт
                prod_id = str(uuid.uuid4()).replace("-", "")
                sku = f"PROD-{pos_code}"
                created_at = datetime.now(timezone.utc).isoformat()
                products_to_insert.append((prod_id, name, float(price), 1, sku, pos_code, category, created_at))
                existing_products_by_code[pos_code] = {"id": prod_id, "name": name, "sku": sku}

        # Выполняем добавление/обновление продуктов
        if products_to_insert:
            cur.executemany("""
                INSERT INTO products (id, name, price, is_active, sku, pos_code, category, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, products_to_insert)
            print(f"  ✅ Добавлено новых продуктов: {len(products_to_insert)}")
        
        if products_to_update:
            cur.executemany("""
                UPDATE products
                SET name = ?, price = ?, category = ?
                WHERE pos_code = ?
            """, products_to_update)
            print(f"  ✅ Обновлено продуктов: {len(products_to_update)}")

        # 3. Читаем рецепты из Excel
        print(f"\n2. Чтение '{RECIPES_XLSX}'...")
        wb_r = openpyxl.load_workbook(RECIPES_XLSX, data_only=True)
        ws_r = wb_r.active

        recipes_to_insert = []
        errors = []
        
        rows_r = list(ws_r.iter_rows(values_only=True))
        
        # Сначала очистим существующие рецепты для импортируемых продуктов,
        # чтобы избежать дублирования или устаревших связей.
        imported_product_ids = set()
        for row in rows_r[1:]:
            if row[0] is None:
                continue
            try:
                pos_code = int(row[0])
                if pos_code in existing_products_by_code:
                    imported_product_ids.add(existing_products_by_code[pos_code]["id"])
            except:
                pass

        if imported_product_ids:
            # Удаляем только технологические карты тех продуктов, которые мы импортируем из Excel
            placeholders = ",".join("?" for _ in imported_product_ids)
            cur.execute(f"DELETE FROM recipes WHERE product_id IN ({placeholders})", list(imported_product_ids))
            print(f"  🧹 Очищены старые технологические карты для {cur.rowcount} записей.")

        created_at_recipe = datetime.now(timezone.utc).isoformat()
        
        for idx, row in enumerate(rows_r[1:], start=2):
            if row[0] is None or str(row[0]).strip() == "":
                continue
            
            try:
                pos_code = int(row[0])
                prod_name = str(row[1]).strip()
                ing_name = str(row[2]).strip()
                quantity = Decimal(str(row[3])) if row[3] is not None else Decimal("0")
            except Exception as e:
                errors.append(f"Строка {idx}: Ошибка парсинга данных ({e})")
                continue

            # Проверяем, существует ли продукт
            if pos_code not in existing_products_by_code:
                errors.append(f"Строка {idx}: Продукт с кодом {pos_code} ('{prod_name}') не найден в базе.")
                continue
            
            prod_id = existing_products_by_code[pos_code]["id"]

            # Ищем ингредиент
            ing_key = ing_name.lower().strip()
            if ing_key not in ingredient_map:
                errors.append(f"Строка {idx}: Ингредиент '{ing_name}' не найден в справочнике сырья!")
                continue
            
            ing_id = ingredient_map[ing_key]
            recipe_id = str(uuid.uuid4()).replace("-", "")

            recipes_to_insert.append((recipe_id, prod_id, ing_id, float(quantity), created_at_recipe))

        # Записываем рецепты
        if recipes_to_insert:
            cur.executemany("""
                INSERT INTO recipes (id, product_id, ingredient_id, quantity, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, recipes_to_insert)
            print(f"  ✅ Загружено связей в рецепты: {len(recipes_to_insert)}")

        conn.commit()
        print("\n=== ИМПОРТ ЗАВЕРШЕН УСПЕШНО ===")
        if errors:
            print("\n⚠️ Предупреждения / Ошибки:")
            for err in errors[:20]:
                print(f"  {err}")
            if len(errors) > 20:
                print(f"  ... и еще {len(errors) - 20} ошибок.")
        else:
            print("  Ошибок нет!")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ Произошла ошибка во время импорта: {e}")
    finally:
        conn.close()

if __name__ == '__main__':
    main()
